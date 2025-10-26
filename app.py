import streamlit as st
import pandas as pd
import numpy as np
import plotly.graph_objects as go
import plotly.express as px
from plotly.subplots import make_subplots
import random
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

# 페이지 설정
st.set_page_config(
    page_title="자리 바꾸기 프로그램",
    page_icon="🏫",
    layout="wide",
    initial_sidebar_state="expanded"
)

# 세션 상태 초기화
if 'students' not in st.session_state:
    st.session_state.students = []
if 'seating_arrangement' not in st.session_state:
    st.session_state.seating_arrangement = {}
if 'pre_assigned_seats' not in st.session_state:
    st.session_state.pre_assigned_seats = {}
if 'disabled_seats' not in st.session_state:
    st.session_state.disabled_seats = []
if 'distanced_students' not in st.session_state:
    st.session_state.distanced_students = []
if 'layout_type' not in st.session_state:
    st.session_state.layout_type = "default"
if 'rows' not in st.session_state:
    st.session_state.rows = 5
if 'cols' not in st.session_state:
    st.session_state.cols = 6
if 'is_teacher_view' not in st.session_state:
    st.session_state.is_teacher_view = False

def initialize_session_state():
    """세션 상태 초기화 함수"""
    if 'students' not in st.session_state:
        st.session_state.students = []
    if 'seating_arrangement' not in st.session_state:
        st.session_state.seating_arrangement = {}
    if 'pre_assigned_seats' not in st.session_state:
        st.session_state.pre_assigned_seats = {}
    if 'disabled_seats' not in st.session_state:
        st.session_state.disabled_seats = []
    if 'distanced_students' not in st.session_state:
        st.session_state.distanced_students = []

def get_seat_coordinates(index, layout_type, rows, cols):
    """자리 인덱스를 좌표로 변환"""
    if layout_type == "pairs":
        # 분단형 배치: 각 분단당 2열씩
        rows_per_section = rows
        desks_per_section = rows_per_section * 2
        section = index // desks_per_section
        index_in_section = index % desks_per_section
        row = index_in_section // 2
        col = (section * 2) + (index_in_section % 2)
        return row, col
    else:
        # 기본 배치
        row = index // cols
        col = index % cols
        return row, col

def is_too_close(index1, index2, layout_type, rows, cols):
    """두 자리가 너무 가까운지 확인"""
    pos1 = get_seat_coordinates(index1, layout_type, rows, cols)
    pos2 = get_seat_coordinates(index2, layout_type, rows, cols)
    
    # 인접한 자리 체크 (상하좌우 + 대각선)
    if abs(pos1[0] - pos2[0]) <= 1 and abs(pos1[1] - pos2[1]) <= 1:
        return True
    # 같은 행에서 2칸 이내
    if pos1[0] == pos2[0] and abs(pos1[1] - pos2[1]) <= 2:
        return True
    # 같은 열에서 2칸 이내
    if pos1[1] == pos2[1] and abs(pos1[0] - pos2[0]) <= 2:
        return True
    return False

def create_seating_chart(seating_arrangement, layout_type, rows, cols, is_teacher_view=False):
    """자리 배치도 생성"""
    if layout_type == "pairs":
        return create_pairs_layout(seating_arrangement, rows, cols, is_teacher_view)
    else:
        return create_default_layout(seating_arrangement, rows, cols, is_teacher_view)

def create_default_layout(seating_arrangement, rows, cols, is_teacher_view=False):
    """기본 격자형 자리 배치도 생성"""
    fig = go.Figure()
    
    # 자리 그리기
    for i in range(rows * cols):
        row = i // cols
        col = i % cols
        
        # 교사 기준 보기일 때 좌표 반전
        if is_teacher_view:
            display_row = rows - 1 - row
            display_col = cols - 1 - col
        else:
            display_row = row
            display_col = col
        
        student_name = seating_arrangement.get(i, "")
        
        # 자리 색상 설정
        if i in st.session_state.disabled_seats:
            color = "lightgray"
            text_color = "gray"
        elif student_name:
            color = "lightblue"
            text_color = "black"
        else:
            color = "white"
            text_color = "gray"
        
        # 자리 사각형 그리기
        fig.add_shape(
            type="rect",
            x0=display_col - 0.4, y0=display_row - 0.4,
            x1=display_col + 0.4, y1=display_row + 0.4,
            fillcolor=color,
            line=dict(color="black", width=2)
        )
        
        # 자리 번호와 학생 이름 표시
        fig.add_annotation(
            x=display_col, y=display_row,
            text=f"{i+1}<br>{student_name}",
            showarrow=False,
            font=dict(size=10, color=text_color),
            xanchor="center",
            yanchor="middle"
        )
    
    # 교탁 표시
    fig.add_shape(
        type="rect",
        x0=-0.5, y0=rows + 0.5,
        x1=cols - 0.5, y1=rows + 1.5,
        fillcolor="lightyellow",
        line=dict(color="black", width=2)
    )
    
    fig.add_annotation(
        x=cols/2 - 0.5, y=rows + 1,
        text="교탁",
        showarrow=False,
        font=dict(size=14, color="black", family="Arial Black"),
        xanchor="center",
        yanchor="middle"
    )
    
    fig.update_layout(
        title="자리 배치도",
        xaxis=dict(
            range=[-1, cols],
            showgrid=True,
            zeroline=False,
            showticklabels=False
        ),
        yaxis=dict(
            range=[-0.5, rows + 2],
            showgrid=True,
            zeroline=False,
            showticklabels=False,
            scaleanchor="x",
            scaleratio=1
        ),
        showlegend=False,
        width=600,
        height=400,
        margin=dict(l=50, r=50, t=50, b=50)
    )
    
    return fig

def create_pairs_layout(seating_arrangement, sections, rows_per_section, is_teacher_view=False):
    """분단형 자리 배치도 생성"""
    fig = go.Figure()
    
    total_desks = sections * rows_per_section * 2
    
    for i in range(total_desks):
        section = i // (rows_per_section * 2)
        index_in_section = i % (rows_per_section * 2)
        row = index_in_section // 2
        col = (section * 2) + (index_in_section % 2)
        
        # 교사 기준 보기일 때 좌표 반전
        if is_teacher_view:
            display_row = rows_per_section - 1 - row
            display_col = (sections * 2) - 1 - col
        else:
            display_row = row
            display_col = col
        
        student_name = seating_arrangement.get(i, "")
        
        # 자리 색상 설정
        if i in st.session_state.disabled_seats:
            color = "lightgray"
            text_color = "gray"
        elif student_name:
            color = "lightblue"
            text_color = "black"
        else:
            color = "white"
            text_color = "gray"
        
        # 자리 사각형 그리기
        fig.add_shape(
            type="rect",
            x0=display_col - 0.4, y0=display_row - 0.4,
            x1=display_col + 0.4, y1=display_row + 0.4,
            fillcolor=color,
            line=dict(color="black", width=2)
        )
        
        # 자리 번호와 학생 이름 표시
        fig.add_annotation(
            x=display_col, y=display_row,
            text=f"{i+1}<br>{student_name}",
            showarrow=False,
            font=dict(size=10, color=text_color),
            xanchor="center",
            yanchor="middle"
        )
    
    # 교탁 표시
    fig.add_shape(
        type="rect",
        x0=-0.5, y0=rows_per_section + 0.5,
        x1=(sections * 2) - 0.5, y1=rows_per_section + 1.5,
        fillcolor="lightyellow",
        line=dict(color="black", width=2)
    )
    
    fig.add_annotation(
        x=(sections * 2)/2 - 0.5, y=rows_per_section + 1,
        text="교탁",
        showarrow=False,
        font=dict(size=14, color="black", family="Arial Black"),
        xanchor="center",
        yanchor="middle"
    )
    
    fig.update_layout(
        title="자리 배치도 (분단형)",
        xaxis=dict(
            range=[-1, sections * 2],
            showgrid=True,
            zeroline=False,
            showticklabels=False
        ),
        yaxis=dict(
            range=[-0.5, rows_per_section + 2],
            showgrid=True,
            zeroline=False,
            showticklabels=False,
            scaleanchor="x",
            scaleratio=1
        ),
        showlegend=False,
        width=600,
        height=400,
        margin=dict(l=50, r=50, t=50, b=50)
    )
    
    return fig

def generate_seating_arrangement():
    """자리 배치 생성"""
    if not st.session_state.students:
        st.error("먼저 학생 명단을 입력해주세요.")
        return
    
    # 랜덤 시드 설정
    random_seed = getattr(st.session_state, 'random_seed', 42)
    random.seed(random_seed)
    np.random.seed(random_seed)
    
    # 총 자리 수 계산
    if st.session_state.layout_type == "pairs":
        total_seats = st.session_state.rows * st.session_state.cols * 2
    else:
        total_seats = st.session_state.rows * st.session_state.cols
    
    # 사용 가능한 자리 계산
    available_seats = [i for i in range(total_seats) 
                      if i not in st.session_state.disabled_seats 
                      and i not in st.session_state.pre_assigned_seats]
    
    if len(available_seats) < len(st.session_state.students):
        st.error(f"사용 가능한 자리({len(available_seats)}개)가 학생 수({len(st.session_state.students)}명)보다 적습니다.")
        return
    
    # 사전 지정된 자리 배치
    final_arrangement = st.session_state.pre_assigned_seats.copy()
    pre_assigned_students = set(st.session_state.pre_assigned_seats.values())
    
    # 자리 띄우기 대상 학생들
    distanced_students = [s for s in st.session_state.distanced_students 
                         if s not in pre_assigned_students]
    
    # 일반 학생들
    regular_students = [s for s in st.session_state.students 
                       if s not in pre_assigned_students and s not in distanced_students]
    
    # 배치 알고리즘 선택
    algorithm = getattr(st.session_state, 'algorithm', '기본')
    
    if algorithm == "균형 배치":
        final_arrangement = generate_balanced_arrangement(
            final_arrangement, distanced_students, regular_students, 
            available_seats, st.session_state.layout_type, 
            st.session_state.rows, st.session_state.cols
        )
    elif algorithm == "그룹 분산":
        final_arrangement = generate_group_distributed_arrangement(
            final_arrangement, distanced_students, regular_students, 
            available_seats, st.session_state.layout_type, 
            st.session_state.rows, st.session_state.cols
        )
    else:
        # 기본 알고리즘
        final_arrangement = generate_default_arrangement(
            final_arrangement, distanced_students, regular_students, 
            available_seats, st.session_state.layout_type, 
            st.session_state.rows, st.session_state.cols
        )
    
    st.session_state.seating_arrangement = final_arrangement
    
    # 자동 히스토리 저장
    auto_save = getattr(st.session_state, 'auto_save', True)
    if auto_save:
        save_to_history(final_arrangement)
    
    st.success("자리 배치가 완료되었습니다!")

def generate_default_arrangement(final_arrangement, distanced_students, regular_students, 
                               available_seats, layout_type, rows, cols):
    """기본 자리 배치 알고리즘"""
    # 자리 띄우기 학생들 배치
    placed_distanced_indices = []
    unplaced_distanced = []
    
    if distanced_students:
        available_for_distanced = [i for i in available_seats 
                                 if i not in final_arrangement]
        random.shuffle(available_for_distanced)
        
        for student in distanced_students:
            placed = False
            for i, seat_index in enumerate(available_for_distanced):
                is_close_to_other = any(is_too_close(seat_index, placed_idx, 
                                                   layout_type, rows, cols) 
                                      for placed_idx in placed_distanced_indices)
                
                if not is_close_to_other:
                    final_arrangement[seat_index] = student
                    placed_distanced_indices.append(seat_index)
                    available_for_distanced.pop(i)
                    placed = True
                    break
            
            if not placed:
                unplaced_distanced.append(student)
    
    # 자리 띄우기에 실패한 학생들을 일반 학생에 추가
    regular_students.extend(unplaced_distanced)
    
    # 일반 학생들 배치
    random.shuffle(regular_students)
    remaining_seats = [i for i in available_seats if i not in final_arrangement]
    
    for i, student in enumerate(regular_students):
        if i < len(remaining_seats):
            final_arrangement[remaining_seats[i]] = student
    
    return final_arrangement

def generate_balanced_arrangement(final_arrangement, distanced_students, regular_students, 
                                available_seats, layout_type, rows, cols):
    """균형 자리 배치 알고리즘 (앞뒤, 좌우 균형 고려)"""
    # 자리 띄우기 학생들 먼저 배치
    placed_distanced_indices = []
    unplaced_distanced = []
    
    if distanced_students:
        available_for_distanced = [i for i in available_seats 
                                 if i not in final_arrangement]
        random.shuffle(available_for_distanced)
        
        for student in distanced_students:
            placed = False
            for i, seat_index in enumerate(available_for_distanced):
                is_close_to_other = any(is_too_close(seat_index, placed_idx, 
                                                   layout_type, rows, cols) 
                                      for placed_idx in placed_distanced_indices)
                
                if not is_close_to_other:
                    final_arrangement[seat_index] = student
                    placed_distanced_indices.append(seat_index)
                    available_for_distanced.pop(i)
                    placed = True
                    break
            
            if not placed:
                unplaced_distanced.append(student)
    
    regular_students.extend(unplaced_distanced)
    
    # 균형 배치를 위한 자리 우선순위 계산
    remaining_seats = [i for i in available_seats if i not in final_arrangement]
    
    # 자리별 가중치 계산 (앞뒤, 좌우 균형 고려)
    seat_weights = {}
    for seat in remaining_seats:
        row, col = get_seat_coordinates(seat, layout_type, rows, cols)
        # 앞쪽과 뒤쪽, 좌측과 우측의 균형을 고려한 가중치
        weight = abs(row - rows/2) + abs(col - cols/2)
        seat_weights[seat] = weight
    
    # 가중치 순으로 자리 정렬
    sorted_seats = sorted(remaining_seats, key=lambda x: seat_weights[x])
    
    # 학생들을 균형있게 배치
    random.shuffle(regular_students)
    for i, student in enumerate(regular_students):
        if i < len(sorted_seats):
            final_arrangement[sorted_seats[i]] = student
    
    return final_arrangement

def generate_group_distributed_arrangement(final_arrangement, distanced_students, regular_students, 
                                         available_seats, layout_type, rows, cols):
    """그룹 분산 자리 배치 알고리즘 (학생들을 여러 그룹으로 나누어 분산 배치)"""
    # 자리 띄우기 학생들 먼저 배치
    placed_distanced_indices = []
    unplaced_distanced = []
    
    if distanced_students:
        available_for_distanced = [i for i in available_seats 
                                 if i not in final_arrangement]
        random.shuffle(available_for_distanced)
        
        for student in distanced_students:
            placed = False
            for i, seat_index in enumerate(available_for_distanced):
                is_close_to_other = any(is_too_close(seat_index, placed_idx, 
                                                   layout_type, rows, cols) 
                                      for placed_idx in placed_distanced_indices)
                
                if not is_close_to_other:
                    final_arrangement[seat_index] = student
                    placed_distanced_indices.append(seat_index)
                    available_for_distanced.pop(i)
                    placed = True
                    break
            
            if not placed:
                unplaced_distanced.append(student)
    
    regular_students.extend(unplaced_distanced)
    
    # 학생들을 그룹으로 나누기
    group_size = max(1, len(regular_students) // 4)  # 4개 그룹으로 나누기
    groups = [regular_students[i:i + group_size] for i in range(0, len(regular_students), group_size)]
    
    remaining_seats = [i for i in available_seats if i not in final_arrangement]
    
    # 각 그룹을 다른 영역에 배치
    seats_per_group = len(remaining_seats) // len(groups) if groups else 0
    
    for group_idx, group in enumerate(groups):
        start_idx = group_idx * seats_per_group
        end_idx = start_idx + seats_per_group if group_idx < len(groups) - 1 else len(remaining_seats)
        group_seats = remaining_seats[start_idx:end_idx]
        
        random.shuffle(group)
        for i, student in enumerate(group):
            if i < len(group_seats):
                final_arrangement[group_seats[i]] = student
    
    return final_arrangement

def save_to_history(arrangement):
    """자리 배치를 히스토리에 저장"""
    if 'seating_history' not in st.session_state:
        st.session_state.seating_history = []
    
    from datetime import datetime
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    history_entry = {
        'timestamp': timestamp,
        'arrangement': arrangement.copy(),
        'students': st.session_state.students.copy(),
        'layout_type': st.session_state.layout_type,
        'rows': st.session_state.rows,
        'cols': st.session_state.cols
    }
    
    st.session_state.seating_history.append(history_entry)
    
    # 히스토리 크기 제한 (최대 20개)
    if len(st.session_state.seating_history) > 20:
        st.session_state.seating_history = st.session_state.seating_history[-20:]

def create_excel_file():
    """엑셀 파일 생성"""
    if not st.session_state.seating_arrangement:
        st.error("먼저 자리 배치를 생성해주세요.")
        return None
    
    # 메모리에서 엑셀 파일 생성
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "자리배치도"
    
    if st.session_state.layout_type == "pairs":
        # 분단형 배치
        sections = st.session_state.rows
        rows_per_section = st.session_state.cols
        
        # 헤더 생성
        header1 = ['']
        header2 = ['행']
        for s in range(sections):
            section_label = f"{sections - s}분단" if st.session_state.is_teacher_view else f"{s + 1}분단"
            header1.extend([section_label, ''])
            header2.extend(['왼쪽', '오른쪽'])
        
        ws.append(header1)
        ws.append(header2)
        
        # 데이터 행 생성
        for r in range(rows_per_section):
            row_label = f"{rows_per_section - r}행" if st.session_state.is_teacher_view else f"{r + 1}행"
            row_data = [row_label]
            
            for s in range(sections):
                read_section = sections - 1 - s if st.session_state.is_teacher_view else s
                read_row = rows_per_section - 1 - r if st.session_state.is_teacher_view else r
                
                student_left_index = (read_section * rows_per_section * 2) + (read_row * 2)
                student_right_index = student_left_index + 1
                
                left_index = student_right_index if st.session_state.is_teacher_view else student_left_index
                right_index = student_left_index if st.session_state.is_teacher_view else student_right_index
                
                left_student = st.session_state.seating_arrangement.get(left_index, "")
                right_student = st.session_state.seating_arrangement.get(right_index, "")
                
                row_data.extend([left_student, right_student])
            
            ws.append(row_data)
        
        # 셀 병합
        for s in range(sections):
            ws.merge_cells(start_row=1, start_column=s*2+2, end_row=1, end_column=s*2+3)
    
    else:
        # 기본 배치
        rows = st.session_state.rows
        cols = st.session_state.cols
        
        # 헤더 생성
        header = [' ']
        for c in range(cols):
            col_label = f"{cols - c}열" if st.session_state.is_teacher_view else f"{c + 1}열"
            header.append(col_label)
        
        ws.append(header)
        
        # 데이터 행 생성
        for r in range(rows):
            row_label = f"{rows - r}행" if st.session_state.is_teacher_view else f"{r + 1}행"
            row_data = [row_label]
            
            for c in range(cols):
                read_row = rows - 1 - r if st.session_state.is_teacher_view else r
                read_col = cols - 1 - c if st.session_state.is_teacher_view else c
                
                index = read_row * cols + read_col
                student = st.session_state.seating_arrangement.get(index, "")
                row_data.append(student)
            
            ws.append(row_data)
    
    # 스타일 적용
    for row in ws.iter_rows():
        for cell in row:
            cell.font = Font(name="맑은 고딕", size=11)
            cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # 파일을 메모리에 저장
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer

# 메인 UI
def main():
    st.title("🏫 자리 바꾸기 프로그램")
    st.markdown("**간편하고 빠른 자리 배치로 교실 분위기를 새롭게! 교실 속 자리 배치 도우미**")
    st.markdown("*Made by 슬쌤 / 📧 seulwhite17@gmail.com*")
    
    # 사이드바
    with st.sidebar:
        st.header("1. 명단 입력")
        
        # 명단 입력
        name_input = st.text_area(
            "학생 이름을 한 줄에 한 명씩 입력하세요",
            height=150,
            placeholder="김자두\n백레몬\n홍석류"
        )
        
        if st.button("명단 적용"):
            students = [name.strip() for name in name_input.split('\n') if name.strip()]
            st.session_state.students = students
            st.success(f"{len(students)}명의 학생이 등록되었습니다.")
        
        # 현재 등록된 학생 수 표시
        if st.session_state.students:
            st.info(f"등록된 학생: {len(st.session_state.students)}명")
            with st.expander("학생 목록 보기"):
                for i, student in enumerate(st.session_state.students, 1):
                    st.write(f"{i}. {student}")
        
        st.header("2. 책상 배열 설정")
        
        # 배치 유형 선택
        layout_type = st.selectbox(
            "배치 유형",
            ["default", "pairs"],
            format_func=lambda x: "기본" if x == "default" else "짝꿍 (분단형)"
        )
        st.session_state.layout_type = layout_type
        
        # 행/열 설정
        if layout_type == "pairs":
            cols = st.number_input("분단 수", min_value=1, max_value=10, value=3)
            rows = st.number_input("행 수", min_value=1, max_value=10, value=5)
        else:
            rows = st.number_input("행 (가로)", min_value=1, max_value=15, value=5)
            cols = st.number_input("열 (세로)", min_value=1, max_value=15, value=6)
        
        st.session_state.rows = rows
        st.session_state.cols = cols
        
        # 배열 적용 버튼
        if st.button("배열 적용"):
            st.session_state.seating_arrangement = {}
            st.session_state.pre_assigned_seats = {}
            st.session_state.disabled_seats = []
            st.success("배치가 적용되었습니다.")
        
        st.header("3. 고급 설정")
        
        # 사전 자리 지정
        with st.expander("🔐 사전 자리 지정"):
            if st.session_state.students:
                selected_student = st.selectbox(
                    "지정할 학생 선택",
                    [""] + st.session_state.students
                )
                
                if selected_student:
                    # 총 자리 수 계산
                    if layout_type == "pairs":
                        total_seats = rows * cols * 2
                    else:
                        total_seats = rows * cols
                    
                    seat_number = st.number_input(
                        "자리 번호",
                        min_value=1,
                        max_value=total_seats,
                        value=1
                    )
                    
                    col1, col2 = st.columns(2)
                    with col1:
                        if st.button("지정"):
                            seat_index = seat_number - 1
                            st.session_state.pre_assigned_seats[seat_index] = selected_student
                            st.success(f"{selected_student} 학생을 {seat_number}번 자리에 지정했습니다.")
                    
                    with col2:
                        if st.button("해제"):
                            seat_index = seat_number - 1
                            if seat_index in st.session_state.pre_assigned_seats:
                                del st.session_state.pre_assigned_seats[seat_index]
                                st.success(f"{seat_number}번 자리 지정이 해제되었습니다.")
                
                # 지정된 자리 목록
                if st.session_state.pre_assigned_seats:
                    st.write("**지정된 자리:**")
                    for seat_idx, student in st.session_state.pre_assigned_seats.items():
                        st.write(f"• {seat_idx + 1}번 자리: {student}")
        
        # 자리 띄우기
        with st.expander("🧍↔️ 자리 띄우기"):
            if st.session_state.students:
                distanced_students = st.multiselect(
                    "서로 붙어 앉으면 안 되는 학생들 선택",
                    st.session_state.students
                )
                st.session_state.distanced_students = distanced_students
                
                if distanced_students:
                    st.info(f"선택된 학생: {', '.join(distanced_students)}")
                    st.caption("💡 안정적인 배치를 위해 전체 자리의 약 1/8 이내 인원을 선택하는 것을 권장합니다.")
        
        # 자리 비활성화
        with st.expander("🚫 자리 비활성화"):
            if st.session_state.students:
                # 총 자리 수 계산
                if layout_type == "pairs":
                    total_seats = rows * cols * 2
                else:
                    total_seats = rows * cols
                
                st.write("**비활성화할 자리 선택:**")
                disabled_seats_input = st.multiselect(
                    "자리 번호 선택 (여러 개 선택 가능)",
                    options=list(range(1, total_seats + 1)),
                    default=[seat + 1 for seat in st.session_state.disabled_seats]
                )
                
                # 비활성화된 자리 업데이트
                st.session_state.disabled_seats = [seat - 1 for seat in disabled_seats_input]
                
                if st.session_state.disabled_seats:
                    st.warning(f"비활성화된 자리: {', '.join(map(str, [seat + 1 for seat in st.session_state.disabled_seats]))}")
                
                if st.button("모든 자리 활성화"):
                    st.session_state.disabled_seats = []
                    st.success("모든 자리가 활성화되었습니다.")
                    st.rerun()
        
        # 배치 히스토리
        with st.expander("📚 배치 히스토리"):
            if 'seating_history' not in st.session_state:
                st.session_state.seating_history = []
            
            if st.session_state.seating_history:
                st.write("**최근 자리 배치 기록:**")
                for i, history in enumerate(st.session_state.seating_history[-5:], 1):  # 최근 5개만 표시
                    with st.container():
                        col1, col2, col3 = st.columns([2, 1, 1])
                        with col1:
                            st.write(f"{i}. {history['timestamp']}")
                        with col2:
                            if st.button(f"불러오기", key=f"load_{i}"):
                                st.session_state.seating_arrangement = history['arrangement']
                                st.success("히스토리가 불러와졌습니다.")
                                st.rerun()
                        with col3:
                            if st.button(f"삭제", key=f"delete_{i}"):
                                st.session_state.seating_history.pop(-i)
                                st.success("히스토리가 삭제되었습니다.")
                                st.rerun()
            else:
                st.info("아직 배치 히스토리가 없습니다.")
            
            if st.button("히스토리 모두 삭제"):
                st.session_state.seating_history = []
                st.success("모든 히스토리가 삭제되었습니다.")
        
        # 배치 통계
        with st.expander("📊 배치 통계"):
            if st.session_state.students:
                # 총 자리 수 계산
                if layout_type == "pairs":
                    total_seats = rows * cols * 2
                else:
                    total_seats = rows * cols
                
                available_seats = total_seats - len(st.session_state.disabled_seats)
                pre_assigned_count = len(st.session_state.pre_assigned_seats)
                distanced_count = len(st.session_state.distanced_students)
                
                st.metric("총 자리 수", total_seats)
                st.metric("사용 가능한 자리", available_seats)
                st.metric("등록된 학생 수", len(st.session_state.students))
                st.metric("사전 지정된 자리", pre_assigned_count)
                st.metric("자리 띄우기 대상", distanced_count)
                
                # 배치 가능성 체크
                if available_seats < len(st.session_state.students):
                    st.error("⚠️ 사용 가능한 자리가 학생 수보다 적습니다!")
                elif available_seats == len(st.session_state.students):
                    st.warning("⚠️ 자리 수와 학생 수가 정확히 일치합니다.")
                else:
                    st.success("✅ 배치 가능합니다.")
        
        # 고급 옵션
        with st.expander("⚙️ 고급 옵션"):
            # 랜덤 시드 설정
            random_seed = st.number_input(
                "랜덤 시드 (재현 가능한 배치를 위해)",
                min_value=0,
                max_value=999999,
                value=42,
                help="같은 시드를 사용하면 동일한 배치 결과를 얻을 수 있습니다."
            )
            
            # 배치 알고리즘 옵션
            algorithm = st.selectbox(
                "배치 알고리즘",
                ["기본", "균형 배치", "그룹 분산"],
                help="다양한 배치 알고리즘을 선택할 수 있습니다."
            )
            
            # 자동 저장 옵션
            auto_save = st.checkbox(
                "자동 히스토리 저장",
                value=True,
                help="자리 배치 생성 시 자동으로 히스토리에 저장합니다."
            )
            
            # 설정 저장
            if st.button("설정 저장"):
                st.session_state.random_seed = random_seed
                st.session_state.algorithm = algorithm
                st.session_state.auto_save = auto_save
                st.success("설정이 저장되었습니다.")
    
    # 메인 영역
    col1, col2 = st.columns([2, 1])
    
    with col1:
        st.header("3. 자리 배치 결과")
        
        # 자리 배치 버튼들
        button_col1, button_col2, button_col3 = st.columns(3)
        
        with button_col1:
            if st.button("🎲 자리 바꾸기!", type="primary", use_container_width=True):
                generate_seating_arrangement()
        
        with button_col2:
            if st.button("🗑️ 모두 지우기", use_container_width=True):
                st.session_state.seating_arrangement = {}
                st.success("모든 자리가 지워졌습니다.")
        
        with button_col3:
            if st.session_state.seating_arrangement:
                excel_buffer = create_excel_file()
                if excel_buffer:
                    st.download_button(
                        label="📊 엑셀로 다운로드",
                        data=excel_buffer.getvalue(),
                        file_name=f"자리배치결과_{'교사기준' if st.session_state.is_teacher_view else '학생기준'}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
        
        # 자리 배치도 표시
        if st.session_state.seating_arrangement:
            fig = create_seating_chart(
                st.session_state.seating_arrangement,
                st.session_state.layout_type,
                st.session_state.rows,
                st.session_state.cols,
                st.session_state.is_teacher_view
            )
            st.plotly_chart(fig, use_container_width=True)
        else:
            st.info("자리 배치를 생성하려면 '자리 바꾸기!' 버튼을 클릭하세요.")
    
    with col2:
        st.header("보기 옵션")
        
        # 교사 기준 보기 토글
        if st.button("👨‍🏫 교사 기준 보기" if not st.session_state.is_teacher_view else "👨‍🎓 학생 기준 보기"):
            st.session_state.is_teacher_view = not st.session_state.is_teacher_view
            st.rerun()
        
        # 사용법 안내
        st.markdown("""
        ### 📖 사용법
        1. **명단 입력**: 학생 이름을 한 줄에 한 명씩 입력
        2. **배치 설정**: 교실 형태와 크기 설정
        3. **고급 설정**: 필요시 사전 지정, 자리 띄우기
        4. **자리 배치**: '자리 바꾸기!' 버튼 클릭
        5. **결과 저장**: 엑셀 파일로 다운로드
        
        ### 💡 팁
        - 자리 띄우기는 전체 자리의 1/8 이내 권장
        - 교사 기준 보기로 교탁에서 보는 시점 확인
        - 사전 지정으로 특별 관리가 필요한 학생 배치
        """)

if __name__ == "__main__":
    initialize_session_state()
    main()
